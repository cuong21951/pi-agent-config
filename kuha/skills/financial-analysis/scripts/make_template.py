#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_template.py - Sinh file mau templates/bctc-mau.xlsx: BCTC minh hoa cho
mot cong ty khach san + cong vien giai tri, 3 ky (2024, 2025, 2026E), don vi
ty dong. So lieu la MINH HOA, dung de nguoi dung sao chep cau truc va thay
bang so lieu that.

Cach dung:
    py -3.12 make_template.py [--out duong_dan_file.xlsx]

Mac dinh ghi ra <thu_muc_skill>/templates/bctc-mau.xlsx
"""
import argparse
import sys
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

PERIODS = ["2024", "2025", "2026E"]

PNL_ROWS = [
    ("Doanh thu thuần", [320, 385, 470]),
    ("Giá vốn hàng bán", [190, 220, 260]),
    ("Lợi nhuận gộp", [130, 165, 210]),
    ("Chi phí bán hàng", [35, 40, 46]),
    ("Chi phí quản lý doanh nghiệp", [28, 32, 36]),
    ("Chi phí lãi vay", [20, 24, 26]),
    ("Lợi nhuận trước thuế", [47, 69, 102]),
    ("Lợi nhuận sau thuế", [37.6, 55.2, 81.6]),
    ("Khấu hao TSCĐ", [45, 52, 58]),
    ("EBITDA", [112, 145, 186]),
]

BS_ROWS = [
    ("Tài sản ngắn hạn", [95, 120, 150]),
    ("  Tiền và tương đương tiền", [28, 35, 45]),
    ("  Phải thu khách hàng", [18, 22, 27]),
    ("  Hàng tồn kho", [12, 15, 18]),
    ("  Tài sản ngắn hạn khác", [37, 48, 60]),
    ("Tài sản dài hạn", [620, 690, 740]),
    ("  Tài sản cố định", [590, 655, 700]),
    ("  Tài sản dài hạn khác", [30, 35, 40]),
    ("Tổng tài sản", [715, 810, 890]),
    ("Nợ ngắn hạn", [110, 130, 135]),
    ("  Phải trả người bán", [25, 30, 33]),
    ("  Vay ngắn hạn", [40, 45, 42]),
    ("  Nợ ngắn hạn khác", [45, 55, 60]),
    ("Nợ dài hạn", [320, 340, 345]),
    ("  Vay dài hạn", [300, 310, 300]),
    ("  Nợ dài hạn khác", [20, 30, 45]),
    ("Nợ phải trả", [430, 470, 480]),
    ("Vốn chủ sở hữu", [285, 340, 410]),
    ("  Vốn góp của chủ sở hữu", [250, 250, 250]),
    ("  Lợi nhuận sau thuế chưa phân phối", [35, 90, 160]),
]


def add_header(ws, row, title):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=12)
    return row + 1


def add_col_header(ws, row):
    ws.cell(row=row, column=1, value="Chỉ tiêu").font = Font(bold=True)
    for i, p in enumerate(PERIODS, start=2):
        c = ws.cell(row=row, column=i, value=p)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
    return row + 1


def add_data_rows(ws, row, rows):
    for label, values in rows:
        ws.cell(row=row, column=1, value=label)
        for i, v in enumerate(values, start=2):
            c = ws.cell(row=row, column=i, value=v)
            c.number_format = "#,##0.0"
        row += 1
    return row


def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "BCTC"

    row = 1
    row = add_header(ws, row, "BÁO CÁO KẾT QUẢ HOẠT ĐỘNG KINH DOANH (tỷ đồng) - SỐ LIỆU MINH HỌA")
    row = add_col_header(ws, row)
    row = add_data_rows(ws, row, PNL_ROWS)
    row += 1

    row = add_header(ws, row, "BẢNG CÂN ĐỐI KẾ TOÁN (tỷ đồng) - SỐ LIỆU MINH HỌA")
    row = add_col_header(ws, row)
    row = add_data_rows(ws, row, BS_ROWS)
    row += 2

    note = (
        "Ghi chú: đây là số liệu MINH HỌA cho công ty khách sạn + công viên giải trí giả định, "
        "dùng để làm mẫu cấu trúc. Hãy thay các giá trị bằng số liệu thật của doanh nghiệp, "
        "giữ nguyên tên chỉ tiêu ở cột A và tên kỳ ở hàng tiêu đề để scripts/ratios.py nhận diện đúng."
    )
    ws.cell(row=row, column=1, value=note).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + len(PERIODS))
    ws.row_dimensions[row].height = 45

    ws.column_dimensions["A"].width = 42
    for i in range(2, 2 + len(PERIODS)):
        ws.column_dimensions[get_column_letter(i)].width = 14

    return wb


def main():
    ap = argparse.ArgumentParser(description="Sinh file mau BCTC minh hoa (xlsx).")
    default_out = Path(__file__).resolve().parent.parent / "templates" / "bctc-mau.xlsx"
    ap.add_argument("--out", default=str(default_out), help="Duong dan file xlsx dau ra")
    args = ap.parse_args()

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = build_workbook()
    wb.save(out_path)
    print(f"Da tao file mau: {out_path.resolve()}")


if __name__ == "__main__":
    main()
