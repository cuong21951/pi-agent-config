#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ratios.py - Tinh cac ty so tai chinh chuan tu file BCTC Excel (Viet Nam).

Cach dung:
    py -3.12 ratios.py <bctc.xlsx> [--sheet TEN_SHEET] [--out report.md] [--charts thu_muc]

Dinh dang file dau vao (mot trong hai kieu, tu dong nhan dien):
  1) Hai cot: Chi tieu | Gia tri            (mot ky bao cao)
  2) Nhieu cot: Chi tieu | 2024 | 2025 | 2026E   (nhieu ky, dong dau la ten ky)

Nhan dien nhan dong (row label) theo kieu mo (fuzzy): khong phan biet
hoa/thuong, khong phan biet dau tieng Viet, cho phep sai lech nho.

Script khong tu suy dien so lieu thieu - neu khong tim thay chi tieu can
thiet cho mot ty so, ty so do se duoc bao cao la "thieu du lieu".
"""
import argparse
import difflib
import re
import sys
import unicodedata
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
except ImportError:
    print("Can cai openpyxl: py -3.12 -m pip install openpyxl", file=sys.stderr)
    raise

# ---------------------------------------------------------------------------
# Chuan hoa nhan (bo dau, ha chu thuong, bo ky tu thua) de so khop mo
# ---------------------------------------------------------------------------

def normalize_label(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = s.replace("đ", "d")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


# ---------------------------------------------------------------------------
# Danh sach chi tieu chuan (canonical) va cac cach ghi thuong gap tren BCTC
# ---------------------------------------------------------------------------

CANON = [
    ("doanh_thu_thuan", [
        "Doanh thu thuần",
        "Doanh thu thuần về bán hàng và cung cấp dịch vụ",
        "Doanh thu bán hàng và cung cấp dịch vụ",
    ]),
    ("gia_von", [
        "Giá vốn hàng bán",
    ]),
    ("loi_nhuan_gop", [
        "Lợi nhuận gộp",
        "Lợi nhuận gộp về bán hàng và cung cấp dịch vụ",
    ]),
    ("chi_phi_lai_vay", [
        "Chi phí lãi vay",
        "Lãi vay phải trả",
    ]),
    ("loi_nhuan_truoc_thue", [
        "Lợi nhuận trước thuế",
        "Tổng lợi nhuận kế toán trước thuế",
        "Lợi nhuận kế toán trước thuế",
    ]),
    ("loi_nhuan_sau_thue", [
        "Lợi nhuận sau thuế",
        "Lợi nhuận sau thuế thu nhập doanh nghiệp",
        "Lợi nhuận sau thuế TNDN",
    ]),
    ("khau_hao", [
        "Khấu hao TSCĐ",
        "Khấu hao và phân bổ",
        "Khấu hao tài sản cố định",
        "Khấu hao",
    ]),
    ("ebitda", [
        "EBITDA",
    ]),
    ("tai_san_ngan_han", [
        "Tài sản ngắn hạn",
    ]),
    ("hang_ton_kho", [
        "Hàng tồn kho",
    ]),
    ("phai_thu_khach_hang", [
        "Phải thu khách hàng",
        "Phải thu ngắn hạn của khách hàng",
    ]),
    ("tong_tai_san", [
        "Tổng tài sản",
        "Tổng cộng tài sản",
    ]),
    ("no_ngan_han", [
        "Nợ ngắn hạn",
    ]),
    ("no_phai_tra", [
        "Nợ phải trả",
        "Tổng nợ phải trả",
    ]),
    ("phai_tra_nguoi_ban", [
        "Phải trả người bán",
        "Phải trả người bán ngắn hạn",
    ]),
    ("von_chu_so_huu", [
        "Vốn chủ sở hữu",
    ]),
]

MATCH_CUTOFF = 0.72


def match_labels(row_labels):
    """Tra ve dict canon_key -> row_index (hoac None), va danh sach canon_key khong tim thay."""
    rows_norm = [normalize_label(r) for r in row_labels]
    used = set()
    found = {}
    missing = []
    for key, aliases in CANON:
        best_idx, best_score = None, 0.0
        for i, rn in enumerate(rows_norm):
            if i in used or not rn:
                continue
            for alias in aliases:
                an = normalize_label(alias)
                if not an:
                    continue
                if an == rn:
                    score = 1.0
                elif an in rn or rn in an:
                    score = 0.9
                else:
                    score = difflib.SequenceMatcher(None, an, rn).ratio()
                if score > best_score:
                    best_score, best_idx = score, i
        if best_idx is not None and best_score >= MATCH_CUTOFF:
            found[key] = best_idx
            used.add(best_idx)
        else:
            found[key] = None
            missing.append(key)
    return found, missing


CANON_VN_NAME = {
    "doanh_thu_thuan": "Doanh thu thuần",
    "gia_von": "Giá vốn hàng bán",
    "loi_nhuan_gop": "Lợi nhuận gộp",
    "chi_phi_lai_vay": "Chi phí lãi vay",
    "loi_nhuan_truoc_thue": "Lợi nhuận trước thuế",
    "loi_nhuan_sau_thue": "Lợi nhuận sau thuế",
    "khau_hao": "Khấu hao TSCĐ",
    "ebitda": "EBITDA",
    "tai_san_ngan_han": "Tài sản ngắn hạn",
    "hang_ton_kho": "Hàng tồn kho",
    "phai_thu_khach_hang": "Phải thu khách hàng",
    "tong_tai_san": "Tổng tài sản",
    "no_ngan_han": "Nợ ngắn hạn",
    "no_phai_tra": "Nợ phải trả",
    "phai_tra_nguoi_ban": "Phải trả người bán",
    "von_chu_so_huu": "Vốn chủ sở hữu",
}


# ---------------------------------------------------------------------------
# Doc file Excel
# ---------------------------------------------------------------------------

def to_number(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace(" ", "")
    if s in ("", "-", "N/A", "n/a"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def load_sheet(path, sheet_name):
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise SystemExit(
                f"Khong tim thay sheet '{sheet_name}'. Cac sheet co san: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    rows = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]
    if not rows:
        raise SystemExit("File rong hoac khong doc duoc du lieu.")

    n_cols = max(len(r) for r in rows)

    def non_empty_count(r):
        return sum(1 for c in r if c is not None and str(c).strip() != "")

    # Dong tieu de la dong dau tien co tu 2 o tro len (bo qua dong tieu de
    # section chi co 1 o, vd "BANG CAN DOI KE TOAN (ty dong)")
    header_idx = 0
    for i, r in enumerate(rows):
        if non_empty_count(r) >= 2:
            header_idx = i
            break

    header = list(rows[header_idx]) + [None] * (n_cols - len(rows[header_idx]))

    # cot dau la nhan chi tieu; cac cot con lai la gia tri theo ky
    period_names = []
    for i, h in enumerate(header[1:], start=1):
        if h is None or str(h).strip() == "":
            period_names.append(f"Cột {i+1}")
        else:
            period_names.append(str(h).strip())

    header_label_norm = normalize_label(header[0])

    labels = []
    data = []  # list các row: list gia tri theo tung ky (cung do dai voi period_names)
    for r in rows[header_idx + 1:]:
        r = list(r) + [None] * (n_cols - len(r))
        label = r[0]
        if label is None or str(label).strip() == "":
            continue
        if normalize_label(label) == header_label_norm:
            continue  # dong tieu de lap lai (vd truoc mot section khac)
        labels.append(str(label).strip())
        data.append([to_number(v) for v in r[1:1 + len(period_names)]])

    return period_names, labels, data


# ---------------------------------------------------------------------------
# Tinh ty so
# ---------------------------------------------------------------------------

def get_val(data, found, key, period_idx):
    idx = found.get(key)
    if idx is None or idx >= len(data):
        return None
    row = data[idx]
    if period_idx >= len(row):
        return None
    return row[period_idx]


def avg_val(data, found, key, period_idx):
    cur = get_val(data, found, key, period_idx)
    if period_idx == 0:
        return cur
    prev = get_val(data, found, key, period_idx - 1)
    if cur is None:
        return None
    if prev is None:
        return cur
    return (cur + prev) / 2.0


def safe_div(a, b):
    if a is None or b is None or b == 0:
        return None
    return a / b


def compute_ratios(period_names, data, found):
    results = []
    for pi, pname in enumerate(period_names):
        g = lambda k: get_val(data, found, k, pi)
        a = lambda k: avg_val(data, found, k, pi)

        tsnh = g("tai_san_ngan_han")
        no_ngan_han = g("no_ngan_han")
        hang_ton_kho = g("hang_ton_kho")
        no_phai_tra = g("no_phai_tra")
        vcsh = g("von_chu_so_huu")
        chi_phi_lai_vay = g("chi_phi_lai_vay")
        lntt = g("loi_nhuan_truoc_thue")
        lnst = g("loi_nhuan_sau_thue")
        doanh_thu = g("doanh_thu_thuan")
        gia_von = g("gia_von")
        loi_nhuan_gop = g("loi_nhuan_gop")
        khau_hao = g("khau_hao")
        ebitda_row = g("ebitda")

        ebit = None
        if lntt is not None and chi_phi_lai_vay is not None:
            ebit = lntt + chi_phi_lai_vay
        elif lntt is not None:
            ebit = lntt

        if ebitda_row is not None:
            ebitda = ebitda_row
            ebitda_note = ""
        elif ebit is not None and khau_hao is not None:
            ebitda = ebit + khau_hao
            ebitda_note = ""
        elif ebit is not None:
            ebitda = ebit
            ebitda_note = " (thiếu khấu hao, tạm lấy EBITDA ~ EBIT)"
        else:
            ebitda = None
            ebitda_note = ""

        r = {
            "period": pname,
            "current_ratio": safe_div(tsnh, no_ngan_han),
            "quick_ratio": safe_div(
                (tsnh - hang_ton_kho) if (tsnh is not None and hang_ton_kho is not None) else None,
                no_ngan_han,
            ),
            "de": safe_div(no_phai_tra, vcsh),
            "interest_coverage": safe_div(ebit, chi_phi_lai_vay),
            "gross_margin": safe_div(loi_nhuan_gop, doanh_thu),
            "ebitda": ebitda,
            "ebitda_margin": safe_div(ebitda, doanh_thu),
            "ebitda_note": ebitda_note,
            "roe": safe_div(lnst, a("von_chu_so_huu")),
            "roa": safe_div(lnst, a("tong_tai_san")),
            "dso": safe_div(a("phai_thu_khach_hang"), doanh_thu),
            "dio": safe_div(a("hang_ton_kho"), gia_von),
            "dpo": safe_div(a("phai_tra_nguoi_ban"), gia_von),
            "asset_turnover": safe_div(doanh_thu, a("tong_tai_san")),
        }
        if r["dso"] is not None:
            r["dso"] *= 365
        if r["dio"] is not None:
            r["dio"] *= 365
        if r["dpo"] is not None:
            r["dpo"] *= 365
        r["ccc"] = None
        if r["dso"] is not None and r["dio"] is not None and r["dpo"] is not None:
            r["ccc"] = r["dso"] + r["dio"] - r["dpo"]

        results.append(r)
    return results


# ---------------------------------------------------------------------------
# Nhan xet tu dong (rule-of-thumb, xem references/vas-ifrs.md)
# ---------------------------------------------------------------------------

RULES = [
    ("current_ratio", "Hệ số thanh toán hiện hành", 1.5, 3.0, "lần"),
    ("quick_ratio", "Hệ số thanh toán nhanh", 0.8, 1.2, "lần"),
    ("de", "Nợ/Vốn chủ sở hữu (D/E)", 0.0, 2.0, "lần"),
    ("interest_coverage", "Hệ số khả năng trả lãi vay", 3.0, None, "lần"),
    ("roe", "ROE", 0.12, 0.20, ""),
    ("roa", "ROA", 0.05, 0.10, ""),
]


def fmt_num(v, pct=False, unit=""):
    if v is None:
        return "thiếu dữ liệu"
    if pct:
        return f"{v * 100:.1f}%"
    return f"{v:.2f}{(' ' + unit) if unit else ''}"


def commentary_for_period(r):
    lines = []
    for key, name, lo, hi, unit in RULES:
        v = r.get(key)
        if v is None:
            continue
        pct = key in ("roe", "roa")
        if lo is not None and v < lo:
            lines.append(
                f"- **{name}** = {fmt_num(v, pct, unit)} — thấp hơn khoảng tham khảo "
                f"({fmt_num(lo, pct, unit)}–{fmt_num(hi, pct, unit) if hi is not None else '∞'})."
            )
        elif hi is not None and v > hi:
            lines.append(
                f"- **{name}** = {fmt_num(v, pct, unit)} — cao hơn khoảng tham khảo "
                f"({fmt_num(lo, pct, unit)}–{fmt_num(hi, pct, unit)})."
            )
    return lines


# ---------------------------------------------------------------------------
# Xuat bao cao Markdown
# ---------------------------------------------------------------------------

RATIO_ROWS = [
    ("current_ratio", "Hệ số thanh toán hiện hành (lần)", False),
    ("quick_ratio", "Hệ số thanh toán nhanh (lần)", False),
    ("de", "Nợ/Vốn chủ sở hữu (lần)", False),
    ("interest_coverage", "Khả năng trả lãi vay (lần)", False),
    ("gross_margin", "Biên lợi nhuận gộp", True),
    ("ebitda_margin", "Biên EBITDA", True),
    ("roe", "ROE", True),
    ("roa", "ROA", True),
    ("dso", "Số ngày phải thu (DSO, ngày)", False),
    ("dio", "Số ngày tồn kho (DIO, ngày)", False),
    ("dpo", "Số ngày phải trả (DPO, ngày)", False),
    ("ccc", "Chu kỳ chuyển đổi tiền mặt (ngày)", False),
    ("asset_turnover", "Vòng quay tổng tài sản (lần)", False),
]


def build_report(period_names, results, found, missing, src_path, chart_files):
    lines = []
    lines.append("# Báo cáo chỉ số tài chính")
    lines.append("")
    lines.append(f"Nguồn dữ liệu: `{src_path}`")
    lines.append(f"Các kỳ báo cáo: {', '.join(period_names)}")
    lines.append("")
    lines.append(
        "Đơn vị chỉ số theo đơn vị nhập liệu trong file nguồn (ví dụ: tỷ đồng). "
        "Các khoảng \"tham khảo\" là thông lệ phổ biến (rule of thumb), không phải chuẩn bắt buộc — "
        "xem `references/vas-ifrs.md`."
    )
    lines.append("")

    lines.append("## Bảng chỉ số theo kỳ")
    lines.append("")
    header = "| Chỉ số | " + " | ".join(period_names) + " |"
    sep = "|---|" + "---|" * len(period_names)
    lines.append(header)
    lines.append(sep)
    for key, name, is_pct in RATIO_ROWS:
        row_vals = []
        for r in results:
            v = r.get(key)
            row_vals.append(fmt_num(v, pct=is_pct))
        lines.append(f"| {name} | " + " | ".join(row_vals) + " |")
    lines.append("")

    ebitda_notes = [r["ebitda_note"] for r in results if r.get("ebitda_note")]
    if ebitda_notes:
        lines.append("_Ghi chú EBITDA:_ " + "; ".join(sorted(set(ebitda_notes))))
        lines.append("")

    lines.append("## Nhận xét tự động")
    lines.append("")
    any_flag = False
    for r in results:
        c = commentary_for_period(r)
        if c:
            any_flag = True
            lines.append(f"**Kỳ {r['period']}:**")
            lines.extend(c)
            lines.append("")
    if not any_flag:
        lines.append("Không có chỉ số nào vượt khoảng tham khảo (rule of thumb) trong dữ liệu tính được.")
        lines.append("")

    lines.append("## Dữ liệu không tìm thấy trong file nguồn")
    lines.append("")
    if missing:
        for key in missing:
            lines.append(f"- {CANON_VN_NAME.get(key, key)}: **thiếu dữ liệu** — không khớp được nhãn dòng nào trong file.")
    else:
        lines.append("Tất cả các chỉ tiêu cần thiết đều tìm thấy trong file nguồn.")
    lines.append("")

    if chart_files:
        lines.append("## Biểu đồ")
        lines.append("")
        for f in chart_files:
            lines.append(f"- `{f}`")
        lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Bieu do (tuy chon)
# ---------------------------------------------------------------------------

def make_charts(period_names, data, found, results, out_dir):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    chart_files = []

    doanh_thu = [get_val(data, found, "doanh_thu_thuan", i) for i in range(len(period_names))]
    gross_margin = [r["gross_margin"] for r in results]
    if any(v is not None for v in doanh_thu):
        fig, ax1 = plt.subplots(figsize=(7, 4.5))
        ax1.bar(period_names, [v if v is not None else 0 for v in doanh_thu], color="#4C78A8", label="Doanh thu thuần")
        ax1.set_ylabel("Doanh thu thuần")
        ax1.set_xlabel("Kỳ")
        ax2 = ax1.twinx()
        ax2.plot(
            period_names,
            [v * 100 if v is not None else None for v in gross_margin],
            color="#E45756", marker="o", label="Biên lợi nhuận gộp (%)",
        )
        ax2.set_ylabel("Biên lợi nhuận gộp (%)")
        fig.suptitle("Xu hướng doanh thu và biên lợi nhuận gộp")
        fig.tight_layout()
        path = out_dir / "doanh_thu_bien_loi_nhuan.png"
        fig.savefig(path, dpi=150)
        plt.close(fig)
        chart_files.append(str(path))

    bar_keys = [("current_ratio", "Thanh toán hiện hành"), ("de", "D/E"), ("roe", "ROE"), ("roa", "ROA")]
    if any(any(r.get(k) is not None for k in [b[0] for b in bar_keys]) for r in results):
        fig, ax = plt.subplots(figsize=(7, 4.5))
        n = len(bar_keys)
        width = 0.8 / n
        x = range(len(period_names))
        for i, (key, label) in enumerate(bar_keys):
            vals = [r.get(key) for r in results]
            vals = [v if v is not None else 0 for v in vals]
            xs = [xi + i * width for xi in x]
            ax.bar(xs, vals, width=width, label=label)
        ax.set_xticks([xi + width * (n - 1) / 2 for xi in x])
        ax.set_xticklabels(period_names)
        ax.set_title("So sánh chỉ số theo kỳ")
        ax.legend()
        fig.tight_layout()
        path = out_dir / "chi_so_theo_ky.png"
        fig.savefig(path, dpi=150)
        plt.close(fig)
        chart_files.append(str(path))

    return chart_files


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Tinh cac ty so tai chinh tu BCTC Excel (Viet Nam).")
    ap.add_argument("input", help="Duong dan file BCTC .xlsx")
    ap.add_argument("--sheet", default=None, help="Ten sheet (mac dinh: sheet dau tien)")
    ap.add_argument("--out", default=None, help="Duong dan file bao cao markdown xuat ra")
    ap.add_argument("--charts", default=None, help="Thu muc luu bieu do PNG (tuy chon)")
    args = ap.parse_args()

    src_path = Path(args.input).resolve()
    if not src_path.exists():
        raise SystemExit(f"Khong tim thay file: {src_path}")

    period_names, labels, data = load_sheet(src_path, args.sheet)
    found, missing = match_labels(labels)
    results = compute_ratios(period_names, data, found)

    print(f"Da doc {len(labels)} dong, {len(period_names)} ky: {', '.join(period_names)}")
    if missing:
        print("Khong tim thay cac chi tieu sau trong file nguon:")
        for key in missing:
            print(f"  - {CANON_VN_NAME.get(key, key)}")
    else:
        print("Da tim thay tat ca chi tieu can thiet.")

    chart_files = []
    if args.charts:
        chart_files = make_charts(period_names, data, found, results, args.charts)
        for f in chart_files:
            print(f"Da luu bieu do: {f}")

    report = build_report(period_names, results, found, missing, src_path, chart_files)

    if args.out:
        out_path = Path(args.out)
        out_path.write_text(report, encoding="utf-8")
        print(f"Da ghi bao cao: {out_path.resolve()}")
    else:
        print(report)


if __name__ == "__main__":
    main()
